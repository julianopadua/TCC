# src/modeling/results_consolidator.py

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

from src.utils import loadConfig, get_path, ensure_dir, list_files, get_logger


EXCLUDED_VARIATIONS = {"gridsearch_smote_weight"}


MODEL_LABELS = {
    "DummyClassifier": "Dummy (baseline)",
    "LogisticRegression": "Regressão Logística",
    "NaiveBayes": "Naive Bayes",
    "RandomForest": "Random Forest",
    "SVMLinear": "SVM (Linear)",
    "XGBoost": "XGBoost",
}

VARIATION_LABELS = {
    "base": "Base (sem SMOTE, sem GridSearch, sem pesos)",
    "gridsearch_smote": "GridSearch + SMOTE",
    "gridsearch_weight": "GridSearch + pesos de classe",
    "prior": "Baseline (prior)",
    "stratified": "Baseline (stratified)",
}


def _invert_modeling_scenarios(cfg: Dict[str, Any]) -> Dict[str, str]:
    """
    cfg['modeling_scenarios']: { scenario_key: folder_name }
    retorna: { folder_name: scenario_key }
    """
    ms = cfg.get("modeling_scenarios", {}) or {}
    inv: Dict[str, str] = {}
    for k, v in ms.items():
        if isinstance(v, str) and v.strip():
            inv[v] = k
    return inv


def _scenario_label_from_folder(folder: str, scenario_key: Optional[str] = None) -> str:
    """
    Gera um nome legível e consistente para o cenário.
    Ex.: base_A_no_rad_knn_calculated -> "Cenário A (sem radiação, KNN, variáveis derivadas)"
    """
    # tenta extrair letra do key (base_A, base_B_calculated, etc.)
    letter = None
    if scenario_key and len(scenario_key) >= 6 and scenario_key.startswith("base_"):
        # base_A ou base_A_calculated
        letter = scenario_key.split("_")[1] if "_" in scenario_key else None

    if not letter:
        # fallback: tenta achar "_A_" etc no folder
        for cand in ["A", "B", "C", "D", "E", "F"]:
            if f"base_{cand.lower()}_" in folder.lower():
                letter = cand
                break

    parts: List[str] = []
    f = folder.lower()

    if "no_rad" in f:
        parts.append("sem radiação")
    elif "with_rad" in f:
        parts.append("com radiação")

    if "knn" in f:
        parts.append("KNN")

    if "drop_rows" in f:
        parts.append("remoção de linhas")

    if "full_original" in f:
        parts.append("base completa")

    if f.endswith("_calculated"):
        parts.append("variáveis derivadas")

    if letter:
        core = f"Cenário {letter}"
    else:
        core = "Cenário"

    if parts:
        return f"{core} ({', '.join(parts)})"
    return core


def _parse_timestamp(ts: Optional[str]) -> pd.Timestamp:
    if not ts:
        return pd.NaT
    return pd.to_datetime(ts, format="%Y%m%d_%H%M%S", errors="coerce")


def _extract_path_context(metrics_file: Path) -> Tuple[str, str, str]:
    """
    Espera algo como: results/<model>/<variation>/<scenario>/metrics_*.json
    Retorna: (model_folder, variation_folder, scenario_folder)
    """
    scenario_folder = metrics_file.parent.name
    variation_folder = metrics_file.parent.parent.name
    model_folder = metrics_file.parent.parent.parent.name
    return model_folder, variation_folder, scenario_folder


def _read_metrics_json(fp: Path) -> Dict[str, Any]:
    with fp.open("r", encoding="utf-8") as f:
        return json.load(f)


def _flatten_one(
    fp: Path,
    scenario_key_by_folder: Dict[str, str],
) -> Dict[str, Any]:
    data = _read_metrics_json(fp)

    model_folder, variation_folder, scenario_folder = _extract_path_context(fp)

    model_type = data.get("model_type") or model_folder
    variation = data.get("variation") or variation_folder
    scenario_folder_json = data.get("scenario") or scenario_folder

    metrics = (data.get("metrics") or {}) if isinstance(data.get("metrics"), dict) else {}
    cm = (metrics.get("confusion_matrix") or {}) if isinstance(metrics.get("confusion_matrix"), dict) else {}
    run_meta = (data.get("run_meta") or {}) if isinstance(data.get("run_meta"), dict) else {}
    settings = (run_meta.get("settings") or {}) if isinstance(run_meta.get("settings"), dict) else {}

    scenario_key = scenario_key_by_folder.get(scenario_folder_json)
    scenario_label = _scenario_label_from_folder(scenario_folder_json, scenario_key=scenario_key)

    variation_desc = data.get("variation_desc")
    variation_label = variation_desc or VARIATION_LABELS.get(variation, variation)

    row = {
        # Identificação
        "scenario_id": scenario_key or "",
        "scenario_folder": scenario_folder_json,
        "scenario_label": scenario_label,
        "model_type": model_type,
        "model_label": MODEL_LABELS.get(model_type, model_type),
        "variation": variation,
        "variation_label": variation_label,
        "timestamp": data.get("timestamp"),
        "timestamp_dt": _parse_timestamp(data.get("timestamp")),

        # Métricas principais (comparação)
        "pr_auc": metrics.get("pr_auc"),
        "roc_auc": metrics.get("roc_auc"),
        "f1": metrics.get("f1"),
        "precision": metrics.get("precision"),
        "recall": metrics.get("recall"),
        "specificity": metrics.get("specificity"),

        # Métricas auxiliares
        "brier_score": metrics.get("brier_score"),
        "accuracy": metrics.get("accuracy"),

        # Confusion matrix
        "tn": cm.get("tn"),
        "fp": cm.get("fp"),
        "fn": cm.get("fn"),
        "tp": cm.get("tp"),

        # Contexto de execução relevante
        "threshold": metrics.get("threshold", run_meta.get("threshold")),
        "cv_splits": settings.get("cv_splits"),
        "scoring": settings.get("scoring"),
        "use_smote": settings.get("use_smote"),
        "feature_scaling": settings.get("feature_scaling"),
        "optimize": settings.get("optimize"),

        # Contexto do dataset (útil academicamente)
        "train_rows": run_meta.get("train_rows"),
        "test_rows": run_meta.get("test_rows"),
        "train_pos_rate": run_meta.get("train_pos_rate"),
        "test_pos_rate": run_meta.get("test_pos_rate"),

        # rastreabilidade
        "source_file": str(fp),
    }

    return row


def _build_df(results_dir: Path) -> pd.DataFrame:
    cfg = loadConfig()
    inv = _invert_modeling_scenarios(cfg)

    all_metric_files = list_files(results_dir, ["metrics_*.json"])
    # filtra variações excluídas
    metric_files = []
    for fp in all_metric_files:
        try:
            _, variation_folder, _ = _extract_path_context(fp)
        except Exception:
            continue
        if variation_folder in EXCLUDED_VARIATIONS:
            continue
        metric_files.append(fp)

    if not metric_files:
        raise FileNotFoundError(
            f"Nenhum metrics_*.json encontrado em {results_dir} após filtros: {EXCLUDED_VARIATIONS}"
        )

    rows = [_flatten_one(fp, inv) for fp in metric_files]
    df = pd.DataFrame(rows)

    # Ordenação consistente
    df = df.sort_values(
        ["scenario_label", "model_label", "variation_label", "timestamp_dt"],
        ascending=[True, True, True, True],
    ).reset_index(drop=True)

    return df


def _latest_per_group(df: pd.DataFrame) -> pd.DataFrame:
    """
    Mantém apenas a execução mais recente por (scenario_folder, model_type, variation).
    """
    if df.empty:
        return df.copy()
    work = df.sort_values("timestamp_dt")
    latest = (
        work.groupby(["scenario_folder", "model_type", "variation"], as_index=False)
        .tail(1)
        .sort_values(["scenario_label", "model_label", "variation_label"])
        .reset_index(drop=True)
    )
    return latest


def _best_by_scenario(df_latest: pd.DataFrame, metric: str) -> pd.DataFrame:
    """
    Melhor por cenário segundo uma métrica.
    """
    if df_latest.empty or metric not in df_latest.columns:
        return pd.DataFrame()
    work = df_latest.dropna(subset=[metric]).copy()
    if work.empty:
        return pd.DataFrame()
    best = (
        work.sort_values(metric, ascending=False)
        .groupby("scenario_folder", as_index=False)
        .head(1)
        .sort_values(["scenario_label", metric], ascending=[True, False])
        .reset_index(drop=True)
    )
    return best


def _academic_view(df: pd.DataFrame) -> pd.DataFrame:
    """
    Seleciona e ordena colunas “acadêmicas”: só o que interessa pra comparar modelos.
    """
    cols = [
        "scenario_id",
        "scenario_label",
        "model_label",
        "variation_label",
        "timestamp_dt",
        "pr_auc",
        "roc_auc",
        "f1",
        "precision",
        "recall",
        "specificity",
        "brier_score",
        "accuracy",
        "tn",
        "fp",
        "fn",
        "tp",
        "threshold",
        "cv_splits",
        "scoring",
        "use_smote",
        "feature_scaling",
        "optimize",
        "train_pos_rate",
        "test_pos_rate",
        "train_rows",
        "test_rows",
    ]
    present = [c for c in cols if c in df.columns]
    out = df[present].copy()

    # deixa timestamp bonitinho (sem timezone, estilo acadêmico)
    if "timestamp_dt" in out.columns:
        out["timestamp_dt"] = pd.to_datetime(out["timestamp_dt"], errors="coerce")

    return out


def _format_excel(path: Path) -> None:
    """
    Ajuste simples (freeze + autofilter + widths).
    Não é obrigatório, mas ajuda bastante na leitura.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # auto width com limite
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells:
                v = cell.value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            width = min(max_len + 2, 48)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, width)

    wb.save(path)


def run_consolidation() -> Tuple[Path, Path]:
    log = get_logger("results.consolidator", kind="results", per_run_file=True)

    modeling_dir = get_path("paths", "data", "modeling")
    results_dir = ensure_dir(Path(modeling_dir) / "results")

    log.info(f"[PATH] results_dir={results_dir}")

    df_all = _build_df(results_dir)
    df_latest = _latest_per_group(df_all)

    # Tabelas “acadêmicas”
    academic_latest = _academic_view(df_latest)
    leaderboard_pr = academic_latest.sort_values("pr_auc", ascending=False).reset_index(drop=True)
    best_pr = _academic_view(_best_by_scenario(df_latest, "pr_auc"))
    best_roc = _academic_view(_best_by_scenario(df_latest, "roc_auc"))
    best_f1 = _academic_view(_best_by_scenario(df_latest, "f1"))

    # Saídas (como você pediu: dentro de results/, fora das pastas de modelo)
    out_csv = results_dir / "results.csv"
    out_xlsx = results_dir / "results.xlsx"

    academic_latest.to_csv(out_csv, index=False, encoding="utf-8-sig")

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        academic_latest.to_excel(writer, sheet_name="academic_latest", index=False)
        leaderboard_pr.to_excel(writer, sheet_name="leaderboard_pr_auc", index=False)
        best_pr.to_excel(writer, sheet_name="best_by_pr_auc", index=False)
        best_roc.to_excel(writer, sheet_name="best_by_roc_auc", index=False)
        best_f1.to_excel(writer, sheet_name="best_by_f1", index=False)

        # sheet extra para auditoria (debug), mas sem poluir a tabela acadêmica
        df_all.to_excel(writer, sheet_name="raw_all_runs", index=False)

    _format_excel(out_xlsx)

    log.info(f"[WRITE] {out_csv}")
    log.info(f"[WRITE] {out_xlsx}")
    log.info(f"[ROWS] all_runs={len(df_all)} latest={len(df_latest)}")

    return out_csv, out_xlsx


def main() -> None:
    run_consolidation()


if __name__ == "__main__":
    main()